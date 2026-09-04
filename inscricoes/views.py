from datetime import datetime
from decimal import Decimal


import hashlib
import hmac
import json
import os

import requests

from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum
from django.http import JsonResponse, HttpResponse, request
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .forms import InscricaoForm
from .models import (
    Inscricao,
    Pagamento,
    ContaPagar,
    ContaReceber,
)
from .pagamentos import criar_preferencia_pagamento


def home(request):
    return render(
        request,
        "inscricoes/home.html",
    )


def nova_inscricao(request):
    """
    Cadastro público do atleta.
    Salva a inscrição e redireciona para o pagamento.
    """

    vagas = (
        180
        - Inscricao.objects.exclude(
            status=Inscricao.CANCELADO
        ).count()
    )

    if vagas <= 0:
        return render(
            request,
            "registration/encerradas.html",
        )

    form = InscricaoForm(
        request.POST or None,
        request.FILES or None,
    )

    if request.method == "POST" and form.is_valid():

        inscricao = form.save(
            commit=False
        )

        inscricao.save()

        return redirect(
            "pagamento",
            numero=inscricao.numero,
        )

    return render(
        request,
        "registration/inscricao.html",
        {
            "form": form,
            "vagas": vagas,
        },
    )


def pagamento(request, numero):
    inscricao = get_object_or_404(
        Inscricao,
        numero=numero,
    )

    pagamento = (
        Pagamento.objects
        .filter(inscricao=inscricao)
        .first()
    )

    if not pagamento:
        try:
            link_pagamento = (
                criar_preferencia_pagamento(inscricao)
            )

        except requests.RequestException as exc:
            return render(
                request,
                "registration/pagamento.html",
                {
                    "inscricao": inscricao,
                    "pagamento": None,
                    "erro_pagamento": (
                        "Não foi possível iniciar o "
                        "pagamento no Mercado Pago."
                    ),
                },
                status=502,
            )

        except RuntimeError as exc:
            return render(
                request,
                "registration/pagamento.html",
                {
                    "inscricao": inscricao,
                    "pagamento": None,
                    "erro_pagamento": str(exc),
                },
                status=500,
            )

        pagamento = Pagamento.objects.create(
            inscricao=inscricao,
            valor=inscricao.valor_total,
            link_pagamento=link_pagamento,
            status=Pagamento.PENDENTE,
        )

    return render(
        request,
        "registration/pagamento.html",
        {
            "inscricao": inscricao,
            "pagamento": pagamento,
        },
    )


def sucesso(request, numero):
    inscricao = get_object_or_404(
        Inscricao,
        numero=numero,
    )

    return render(
        request,
        "registration/sucesso.html",
        {
            "inscricao": inscricao,
        },
    )


@staff_member_required
def dashboard(request):
    """
    Dashboard principal do evento.
    Consolida inscrições, pagamentos, receitas extras,
    contas a pagar e indicadores.
    """

    hoje = timezone.localdate()

    # =====================================================
    # INSCRIÇÕES
    # =====================================================

    inscritos = (
        Inscricao.objects
        .exclude(
            status=Inscricao.CANCELADO
        )
    )

    total_inscricoes = inscritos.count()

    total_canceladas = (
        Inscricao.objects
        .filter(
            status=Inscricao.CANCELADO
        )
        .count()
    )

    # =====================================================
    # PAGAMENTOS DAS INSCRIÇÕES
    # =====================================================

    pagamentos_pagos = (
        Pagamento.objects
        .filter(
            status=Pagamento.PAGO
        )
    )

    pagamentos_pendentes = (
        Pagamento.objects
        .filter(
            status=Pagamento.PENDENTE
        )
    )

    total_pagas = pagamentos_pagos.count()
    total_pendentes = pagamentos_pendentes.count()

    recebido_pagamentos = (
        pagamentos_pagos
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    a_receber_pagamentos = (
        pagamentos_pendentes
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    # =====================================================
    # CONTAS A RECEBER / RECEITAS EXTRAS
    # =====================================================

    receitas_recebidas = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.RECEBIDO
        )
    )

    receitas_pendentes = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.PENDENTE
        )
    )

    receitas_vencidas = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.PENDENTE,
            vencimento__lt=hoje,
        )
    )

    recebido_extras = (
        receitas_recebidas
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    a_receber_extras = (
        receitas_pendentes
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    receitas_extras_vencidas = (
        receitas_vencidas
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    # =====================================================
    # TOTAL DE RECEITAS
    # =====================================================

    recebido = (
        recebido_pagamentos
        + recebido_extras
    )

    a_receber = (
        a_receber_pagamentos
        + a_receber_extras
    )

    receita_prevista = (
        recebido
        + a_receber
    )

    # =====================================================
    # CONTAS A PAGAR
    # =====================================================

    despesas_pagas = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.PAGO
        )
    )

    despesas_pendentes = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.PENDENTE
        )
    )

    despesas_vencidas = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.PENDENTE,
            vencimento__lt=hoje,
        )
    )

    contas_pagas = (
        despesas_pagas
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    contas_pendentes = (
        despesas_pendentes
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    contas_vencidas = (
        despesas_vencidas
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    total_despesas = (
        contas_pagas
        + contas_pendentes
    )

    # =====================================================
    # RESULTADO FINANCEIRO
    # =====================================================

    saldo_atual = (
        recebido
        - contas_pagas
    )

    resultado_previsto = (
        receita_prevista
        - total_despesas
    )

    # =====================================================
    # VAGAS
    # =====================================================

    vagas_totais = 180

    vagas_restantes = max(
        vagas_totais - total_inscricoes,
        0,
    )

    ocupacao = (
        total_inscricoes
        / vagas_totais
        * 100
        if vagas_totais
        else 0
    )

    # =====================================================
    # MODALIDADES
    # =====================================================

    mini_sprint = (
        inscritos
        .filter(
            modalidade=Inscricao.MINI
        )
        .count()
    )

    sprint = (
        inscritos
        .filter(
            modalidade=Inscricao.SPRINT
        )
        .count()
    )

    # =====================================================
    # PERFIL DOS PARTICIPANTES
    # =====================================================

    menor_17 = 0
    idade_17_59 = 0
    idade_60_mais = 0
    total_militares = 0

    for atleta in inscritos.iterator():

        idade = atleta.idade_no_evento

        if idade is None:
            continue

        if idade < 17:
            menor_17 += 1
        elif idade <= 59:
            idade_17_59 += 1
        else:
            idade_60_mais += 1

        if atleta.militar:
            total_militares += 1

    percentual_pagamentos = (
        total_pagas
        / total_inscricoes
        * 100
        if total_inscricoes
        else 0
    )

    # =====================================================
    # ÚLTIMOS LANÇAMENTOS
    # =====================================================

    ultimas_contas_pagar = (
        ContaPagar.objects
        .select_related("fornecedor")
        .order_by("-criado_em")[:5]
    )

    ultimas_contas_receber = (
        ContaReceber.objects
        .order_by("-criado_em")[:5]
    )

    # =====================================================
    # CONTEXTO
    # =====================================================

    contexto = {
        "total_inscricoes": total_inscricoes,
        "total_canceladas": total_canceladas,
        "total_pagas": total_pagas,
        "total_pendentes": total_pendentes,
        "recebido": recebido,
        "a_receber": a_receber,
        "receita_prevista": receita_prevista,
        "receitas_extras_vencidas": receitas_extras_vencidas,
        "contas_pagas": contas_pagas,
        "contas_pendentes": contas_pendentes,
        "contas_vencidas": contas_vencidas,
        "total_despesas": total_despesas,
        "saldo_atual": saldo_atual,
        "resultado_previsto": resultado_previsto,
        "vagas_totais": vagas_totais,
        "vagas_restantes": vagas_restantes,
        "ocupacao": round(ocupacao, 1),
        "mini_sprint": mini_sprint,
        "sprint": sprint,
        "menor_17": menor_17,
        "idade_17_59": idade_17_59,
        "idade_60_mais": idade_60_mais,
        "total_militares": total_militares,
        "percentual_pagamentos": round(
            percentual_pagamentos,
            1,
        ),
        "ultimas_contas_pagar": ultimas_contas_pagar,
        "ultimas_contas_receber": ultimas_contas_receber,
    }
   

    return render(
        request,
        "inscricoes/dashboard.html",
        contexto,
    )


@staff_member_required
def extrato_financeiro(request):
    """
    Extrato financeiro por período.

    Considera somente movimentações efetivamente realizadas:

    RECEITAS:
    - Pagamentos de inscrições aprovados
    - Contas a receber com status RECEBIDO

    DESPESAS:
    - Contas a pagar com status PAGO
    """

    hoje = timezone.localdate()

    # =====================================================
    # PERÍODO
    # =====================================================

    inicio_str = request.GET.get("inicio")
    fim_str = request.GET.get("fim")

    if inicio_str:
        try:
            inicio = datetime.strptime(
                inicio_str,
                "%Y-%m-%d",
            ).date()
        except ValueError:
            inicio = hoje.replace(day=1)
    else:
        inicio = hoje.replace(day=1)

    if fim_str:
        try:
            fim = datetime.strptime(
                fim_str,
                "%Y-%m-%d",
            ).date()
        except ValueError:
            fim = hoje
    else:
        fim = hoje

    # Evita período invertido
    if inicio > fim:
        inicio, fim = fim, inicio

    # =====================================================
    # LANÇAMENTOS
    # =====================================================

    lancamentos = []

    # =====================================================
    # PAGAMENTOS DAS INSCRIÇÕES
    # =====================================================

    pagamentos = (
        Pagamento.objects
        .select_related("inscricao")
        .filter(
            status=Pagamento.PAGO,
            pago_em__date__gte=inicio,
            pago_em__date__lte=fim,
        )
        .order_by("pago_em")
    )

    for pagamento in pagamentos:

        lancamentos.append({
            "data": pagamento.pago_em.date(),
            "descricao": (
                f"Inscrição #{pagamento.inscricao.numero} - "
                f"{pagamento.inscricao.nome}"
            ),
            "categoria": "Inscrição",
            "tipo": "RECEITA",
            "valor": pagamento.valor,
        })

    # =====================================================
    # CONTAS A RECEBER
    # =====================================================

    receitas = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.RECEBIDO,
            recebido_em__gte=inicio,
            recebido_em__lte=fim,
        )
        .order_by("recebido_em")
    )

    for conta in receitas:

        lancamentos.append({
            "data": conta.recebido_em,
            "descricao": conta.descricao,
            "categoria": conta.categoria,
            "tipo": "RECEITA",
            "valor": conta.valor,
        })

    # =====================================================
    # CONTAS A PAGAR
    # =====================================================

    despesas = (
        ContaPagar.objects
        .select_related("fornecedor")
        .filter(
            status=ContaPagar.PAGO,
            pago_em__gte=inicio,
            pago_em__lte=fim,
        )
        .order_by("pago_em")
    )

    for conta in despesas:

        fornecedor = ""

        if conta.fornecedor:
            fornecedor = conta.fornecedor.nome

        descricao = conta.descricao

        if fornecedor:
            descricao = f"{fornecedor} - {descricao}"

        lancamentos.append({
            "data": conta.pago_em,
            "descricao": descricao,
            "categoria": conta.categoria,
            "tipo": "DESPESA",
            "valor": conta.valor,
        })

    # =====================================================
    # ORDENAÇÃO
    # =====================================================

    lancamentos.sort(
        key=lambda item: item["data"]
    )

    # =====================================================
    # TOTAIS
    # =====================================================

    total_receitas = sum(
        item["valor"]
        for item in lancamentos
        if item["tipo"] == "RECEITA"
    )

    total_despesas = sum(
        item["valor"]
        for item in lancamentos
        if item["tipo"] == "DESPESA"
    )

    resultado = (
        total_receitas
        - total_despesas
    )

    # =====================================================
    # CONTEXTO
    # =====================================================

    contexto = {
        "inicio": inicio,
        "fim": fim,

        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "resultado": resultado,

        "lancamentos": lancamentos,
    }

    return render(
        request,
        "inscricoes/extrato.html",
        contexto,
    )


@staff_member_required
def relatorios(request):
    """
    Relatório geral do evento.
    Consolida inscrições, pagamentos,
    contas a receber e contas a pagar.
    """

    hoje = timezone.localdate()

    # =====================================================
    # INSCRIÇÕES
    # =====================================================

    total_inscricoes = (
        Inscricao.objects
        .exclude(
            status=Inscricao.CANCELADO
        )
        .count()
    )

    # =====================================================
    # PAGAMENTOS DAS INSCRIÇÕES
    # =====================================================

    total_pagas = (
        Pagamento.objects
        .filter(
            status=Pagamento.PAGO
        )
        .count()
    )

    total_pendentes = (
        Pagamento.objects
        .filter(
            status=Pagamento.PENDENTE
        )
        .count()
    )

    total_recebido_inscricoes = (
        Pagamento.objects
        .filter(
            status=Pagamento.PAGO
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    total_a_receber_inscricoes = (
        Pagamento.objects
        .filter(
            status=Pagamento.PENDENTE
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    # =====================================================
    # CONTAS A RECEBER
    # =====================================================

    total_recebido_extras = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.RECEBIDO
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    total_a_receber_extras = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.PENDENTE
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    total_receitas_extras_vencidas = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.PENDENTE,
            vencimento__lt=hoje
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    # =====================================================
    # TOTAL DE RECEITAS
    # =====================================================

    total_recebido = (
        total_recebido_inscricoes
        + total_recebido_extras
    )

    total_a_receber = (
        total_a_receber_inscricoes
        + total_a_receber_extras
    )

    receita_prevista = (
        total_recebido
        + total_a_receber
    )

    # =====================================================
    # CONTAS A PAGAR
    # =====================================================

    total_contas_pagas = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.PAGO
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    total_contas_pendentes = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.PENDENTE
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    total_contas_vencidas = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.VENCIDO
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    # Também considera pendentes já vencidas,
    # mesmo que o status ainda esteja PENDENTE.
    total_pendentes_vencidas = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.PENDENTE,
            vencimento__lt=hoje
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    despesas_previstas = (
        total_contas_pagas
        + total_contas_pendentes
    )

    # =====================================================
    # RESULTADO
    # =====================================================

    saldo_atual = (
        total_recebido
        - total_contas_pagas
    )

    resultado_previsto = (
        receita_prevista
        - despesas_previstas
    )

    # =====================================================
    # MODALIDADES
    # =====================================================

    mini_sprint = (
        Inscricao.objects
        .filter(
            modalidade=Inscricao.MINI
        )
        .exclude(
            status=Inscricao.CANCELADO
        )
        .count()
    )

    sprint = (
        Inscricao.objects
        .filter(
            modalidade=Inscricao.SPRINT
        )
        .exclude(
            status=Inscricao.CANCELADO
        )
        .count()
    )

    # =====================================================
    # PERCENTUAL DE PAGAMENTOS
    # =====================================================

    percentual_pagamentos = (
        total_pagas
        / total_inscricoes
        * 100
        if total_inscricoes
        else 0
    )

    # =====================================================
    # CONTEXTO
    # =====================================================

    contexto = {
        "hoje": hoje,

        "total_inscricoes":
            total_inscricoes,

        "total_pagas":
            total_pagas,

        "total_pendentes":
            total_pendentes,

        "total_recebido":
            total_recebido,

        "total_a_receber":
            total_a_receber,

        "receita_prevista":
            receita_prevista,

        "total_recebido_inscricoes":
            total_recebido_inscricoes,

        "total_recebido_extras":
            total_recebido_extras,

        "total_a_receber_inscricoes":
            total_a_receber_inscricoes,

        "total_a_receber_extras":
            total_a_receber_extras,

        "total_receitas_extras_vencidas":
            total_receitas_extras_vencidas,

        "total_contas_pagas":
            total_contas_pagas,

        "total_contas_pendentes":
            total_contas_pendentes,

        "total_contas_vencidas":
            total_contas_vencidas,

        "total_pendentes_vencidas":
            total_pendentes_vencidas,

        "despesas_previstas":
            despesas_previstas,

        "saldo_atual":
            saldo_atual,

        "resultado_previsto":
            resultado_previsto,

        "mini_sprint":
            mini_sprint,

        "sprint":
            sprint,

        "percentual_pagamentos":
            round(
                percentual_pagamentos,
                1
            ),
    }

    return render(
        request,
        "inscricoes/relatorio.html",
        contexto,
    )

@csrf_exempt
def webhook_mercadopago(request):
    """
    Recebe notificações de pagamento do Mercado Pago
    e atualiza automaticamente a inscrição correspondente.
    """

    if request.method != "POST":
        return JsonResponse(
            {"detail": "Método não permitido"},
            status=405,
        )

    try:
        body = json.loads(
            request.body.decode("utf-8") or "{}"
        )
    except (
        json.JSONDecodeError,
        UnicodeDecodeError,
    ):
        body = {}

    tipo = (
        request.GET.get("type")
        or body.get("type")
    )

    payment_id = (
        request.GET.get("data.id")
        or body.get("data", {}).get("id")
    )

    # Outros eventos não interessam neste endpoint.
    if tipo and tipo != "payment":
        return JsonResponse(
            {"received": True}
        )

    if not payment_id:
        return JsonResponse(
            {
                "received": True,
                "detail": "Notificação sem payment_id.",
            }
        )

    payment_id = str(payment_id)

    # ==================================================
    # VALIDAR ASSINATURA DO WEBHOOK
    # ==================================================

    secret = os.getenv(
        "MERCADOPAGO_WEBHOOK_SECRET"
    )

    if secret:
        x_signature = request.headers.get(
            "x-signature",
            "",
        )

        x_request_id = request.headers.get(
            "x-request-id",
            "",
        )

        ts = None
        v1 = None

        for item in x_signature.split(","):
            key, _, value = item.strip().partition("=")

            if key == "ts":
                ts = value

            elif key == "v1":
                v1 = value

        if not ts or not v1:
            return JsonResponse(
                {"detail": "Assinatura inválida."},
                status=401,
            )

        manifest = (
            f"id:{payment_id};"
            f"request-id:{x_request_id};"
            f"ts:{ts};"
        )

        generated = hmac.new(
            secret.encode("utf-8"),
            manifest.encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()

        if not hmac.compare_digest(
            generated,
            v1,
        ):
            return JsonResponse(
                {"detail": "Assinatura inválida."},
                status=401,
            )

    # ==================================================
    # CONSULTAR PAGAMENTO NA API DO MERCADO PAGO
    # ==================================================

    access_token = os.getenv(
        "MERCADOPAGO_ACCESS_TOKEN"
    )

    if not access_token:
        return JsonResponse(
            {
                "detail":
                    "MERCADOPAGO_ACCESS_TOKEN não configurado.",
            },
            status=500,
        )

    try:
        response = requests.get(
            (
                "https://api.mercadopago.com/"
                f"v1/payments/{payment_id}"
            ),
            headers={
                "Authorization":
                    f"Bearer {access_token}",
            },
            timeout=20,
        )

        response.raise_for_status()
        data = response.json()

    except requests.RequestException:
        return JsonResponse(
            {
                "detail":
                    "Falha ao consultar o Mercado Pago.",
            },
            status=502,
        )

    # ==================================================
    # IDENTIFICAR INSCRIÇÃO
    # ==================================================

    external_reference = data.get(
        "external_reference"
    )

    if not external_reference:
        return JsonResponse(
            {
                "received": True,
                "payment_id": payment_id,
                "detail":
                    "Pagamento sem external_reference.",
            }
        )

    try:
        inscricao = Inscricao.objects.get(
            numero=external_reference
        )

    except Inscricao.DoesNotExist:
        return JsonResponse(
            {
                "received": True,
                "payment_id": payment_id,
                "external_reference":
                    external_reference,
                "detail":
                    "Inscrição não encontrada.",
            }
        )

    # ==================================================
    # PAGAMENTO LOCAL
    # ==================================================

    pagamento, _ = (
        Pagamento.objects.get_or_create(
            inscricao=inscricao,
            defaults={
                "valor": inscricao.valor_total,
                "link_pagamento": (
                    "https://www.mercadopago.com.br/"
                ),
                "status": Pagamento.PENDENTE,
            },
        )
    )

    # ==================================================
    # VALIDAÇÃO DO VALOR
    # ==================================================

    valor_mp = data.get("transaction_amount")

    try:
        valor_mp = Decimal(str(valor_mp))
    except Exception:
        valor_mp = None

    if (
        valor_mp is not None
        and valor_mp != pagamento.valor
    ):
        return JsonResponse(
            {
                "received": True,
                "payment_id": payment_id,
                "detail":
                    "Valor do pagamento divergente.",
            },
            status=400,
        )

    # ==================================================
    # ATUALIZAÇÃO
    # ==================================================

    status_mp = data.get("status")

    pagamento.identificador_transacao = (
        payment_id
    )

    pagamento.metodo = (
        data.get("payment_method_id")
        or data.get("payment_type_id")
        or ""
    )

    if status_mp == "approved":

        pagamento.status = Pagamento.PAGO

        if not pagamento.pago_em:
            pagamento.pago_em = timezone.now()

        if inscricao.status != Inscricao.PAGO:
            inscricao.status = Inscricao.PAGO

            inscricao.save(
                update_fields=[
                    "status",
                    "atualizado_em",
                ]
            )

    elif status_mp in (
        "cancelled",
        "rejected",
    ):
        pagamento.status = Pagamento.CANCELADO

    elif status_mp == "expired":
        pagamento.status = Pagamento.EXPIRADO

    else:
        pagamento.status = Pagamento.PENDENTE

    pagamento.save()

    return JsonResponse(
        {
            "received": True,
            "payment_id": payment_id,
            "external_reference":
                external_reference,
            "status": status_mp,
        }
    )
@staff_member_required
def exportar_excel(request):

    hoje = timezone.localdate()

    # -------------------------------------------------
    # DADOS
    # -------------------------------------------------

    total_inscricoes = (
        Inscricao.objects
        .exclude(
            status=Inscricao.CANCELADO
        )
        .count()
    )

    total_pagas = (
        Pagamento.objects
        .filter(
            status=Pagamento.PAGO
        )
        .count()
    )

    total_pendentes = (
        Pagamento.objects
        .filter(
            status=Pagamento.PENDENTE
        )
        .count()
    )

    recebido_inscricoes = (
        Pagamento.objects
        .filter(
            status=Pagamento.PAGO
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    a_receber_inscricoes = (
        Pagamento.objects
        .filter(
            status=Pagamento.PENDENTE
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    recebido_extras = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.RECEBIDO
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    a_receber_extras = (
        ContaReceber.objects
        .filter(
            status=ContaReceber.PENDENTE
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    contas_pagas = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.PAGO
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    contas_pendentes = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.PENDENTE
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    contas_vencidas = (
        ContaPagar.objects
        .filter(
            status=ContaPagar.VENCIDO
        )
        .aggregate(
            total=Sum("valor")
        )["total"]
        or 0
    )

    total_recebido = (
        recebido_inscricoes
        + recebido_extras
    )

    total_a_receber = (
        a_receber_inscricoes
        + a_receber_extras
    )

    receita_prevista = (
        total_recebido
        + total_a_receber
    )

    despesas_previstas = (
        contas_pagas
        + contas_pendentes
        + contas_vencidas
    )

    saldo_atual = (
        total_recebido
        - contas_pagas
    )

    resultado_previsto = (
        receita_prevista
        - despesas_previstas
    )

    # -------------------------------------------------
    # CRIA PLANILHA
    # -------------------------------------------------

    workbook = Workbook()

    resumo = workbook.active
    resumo.title = "Resumo"

    resumo["A1"] = "FEST AQUATHLON 2026"
    resumo["A1"].font = Font(
        bold=True,
        size=16,
    )

    resumo["A2"] = "Relatório financeiro"
    resumo["A2"].font = Font(
        bold=True,
    )

    dados = [
        ("Data do relatório", hoje),
        ("Total de inscrições", total_inscricoes),
        ("Pagamentos pagos", total_pagas),
        ("Pagamentos pendentes", total_pendentes),
        ("Total recebido", float(total_recebido)),
        ("Total a receber", float(total_a_receber)),
        ("Receita prevista", float(receita_prevista)),
        ("Contas pagas", float(contas_pagas)),
        ("Contas pendentes", float(contas_pendentes)),
        ("Contas vencidas", float(contas_vencidas)),
        ("Despesas previstas", float(despesas_previstas)),
        ("Saldo atual", float(saldo_atual)),
        ("Resultado previsto", float(resultado_previsto)),
    ]

    linha = 4

    for descricao, valor in dados:

        resumo.cell(
            row=linha,
            column=1,
            value=descricao,
        )

        resumo.cell(
            row=linha,
            column=2,
            value=valor,
        )

        linha += 1

    resumo.column_dimensions["A"].width = 30
    resumo.column_dimensions["B"].width = 20

    # -------------------------------------------------
    # ABA INSCRIÇÕES
    # -------------------------------------------------

    aba_inscricoes = workbook.create_sheet(
        "Inscrições"
    )

    cabecalho = [
        "Número",
        "Nome",
        "Telefone",
        "E-mail",
        "Nascimento",
        "Idade",
        "Modalidade",
        "Militar",
        "Lote",
        "Valor",
        "Status",
    ]

    aba_inscricoes.append(cabecalho)

    for celula in aba_inscricoes[1]:
        celula.font = Font(bold=True)

    atletas = (
        Inscricao.objects
        .all()
        .order_by("numero")
    )

    for atleta in atletas:

        aba_inscricoes.append([
            atleta.numero,
            atleta.nome,
            atleta.telefone,
            atleta.email,
            atleta.data_nascimento,
            atleta.idade_no_evento,
            atleta.get_modalidade_display(),
            "Sim" if atleta.militar else "Não",
            atleta.lote,
            float(atleta.valor_total),
            atleta.get_status_display(),
        ])

    # -------------------------------------------------
    # ABA CONTAS A RECEBER
    # -------------------------------------------------

    aba_receber = workbook.create_sheet(
        "Contas a Receber"
    )

    aba_receber.append([
        "Descrição",
        "Categoria",
        "Valor",
        "Vencimento",
        "Status",
        "Recebido em",
    ])

    for celula in aba_receber[1]:
        celula.font = Font(bold=True)

    for conta in ContaReceber.objects.all():

        aba_receber.append([
            conta.descricao,
            conta.categoria,
            float(conta.valor),
            conta.vencimento,
            conta.get_status_display(),
            conta.recebido_em,
        ])

    # -------------------------------------------------
    # ABA CONTAS A PAGAR
    # -------------------------------------------------

    aba_pagar = workbook.create_sheet(
        "Contas a Pagar"
    )

    aba_pagar.append([
        "Fornecedor",
        "Descrição",
        "Categoria",
        "Valor",
        "Vencimento",
        "Status",
        "Pago em",
    ])

    for celula in aba_pagar[1]:
        celula.font = Font(bold=True)

    for conta in ContaPagar.objects.select_related(
        "fornecedor"
    ).all():

        aba_pagar.append([
            conta.fornecedor.nome,
            conta.descricao,
            conta.categoria,
            float(conta.valor),
            conta.vencimento,
            conta.get_status_display(),
            conta.pago_em,
        ])

    # -------------------------------------------------
    # ABA PAGAMENTOS
    # -------------------------------------------------

    aba_pagamentos = workbook.create_sheet(
        "Pagamentos"
    )

    aba_pagamentos.append([
        "Inscrição",
        "Atleta",
        "Valor",
        "Status",
        "Método",
        "Transação",
        "Criado em",
        "Pago em",
    ])

    for celula in aba_pagamentos[1]:
        celula.font = Font(bold=True)

    for pagamento in Pagamento.objects.select_related(
        "inscricao"
    ).all():

        aba_pagamentos.append([
            pagamento.inscricao.numero,
            pagamento.inscricao.nome,
            float(pagamento.valor),
            pagamento.get_status_display(),
            pagamento.metodo,
            pagamento.identificador_transacao,
            pagamento.criado_em,
            pagamento.pago_em,
        ])

    # -------------------------------------------------
    # FORMATAÇÃO
    # -------------------------------------------------

    for sheet in workbook.worksheets:

        for coluna in sheet.columns:

            tamanho = 0

            for celula in coluna:

                valor = (
                    ""
                    if celula.value is None
                    else str(celula.value)
                )

                tamanho = max(
                    tamanho,
                    len(valor),
                )

            letra = coluna[0].column_letter

            sheet.column_dimensions[
                letra
            ].width = min(
                tamanho + 3,
                45,
            )

        for row in sheet.iter_rows():

            for celula in row:

                celula.alignment = Alignment(
                    vertical="center"
                )

    # -------------------------------------------------
    # DOWNLOAD
    # -------------------------------------------------

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="fest_aquathlon_relatorio.xlsx"'
    )

    workbook.save(response)

    return response